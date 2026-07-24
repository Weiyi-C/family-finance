"""微信账单解析器（CSV 和 Excel）"""

import csv
from datetime import datetime

from .utils import identify_platform_and_merchant


def parse_wechat_csv(content: str) -> tuple[list[dict], dict]:
    """解析微信 CSV 账单

    Returns:
        (items, meta)
    """
    lines = content.strip().split("\n")
    header_idx = -1
    for i, line in enumerate(lines):
        if "交易时间" in line and "交易对方" in line:
            header_idx = i
            break
    if header_idx == -1:
        raise ValueError("无法识别微信账单格式")

    reader = csv.DictReader(lines[header_idx:])
    items = []
    methods = set()

    for row in reader:
        try:
            amount_str = row.get("金额(元)", "0").replace(",", "").replace("¥", "").strip()
            amount = float(amount_str)
            direction = row.get("收/支", "").strip()
            if "不计收支" in direction:
                continue
            txn_type = "expense" if "支出" in direction else "income"
            status = row.get("当前状态", "").strip()
            if "已退款" in status or "退款" in status:
                continue

            # 微信账单有明确的"支付方式"列
            payment_method = row.get("支付方式", "").strip()
            if payment_method:
                methods.add(payment_method)

            # 智能识别平台和商户
            merchant = row.get("交易对方", "").strip()
            description = row.get("商品", "").strip()
            detected_platform, detected_merchant = identify_platform_and_merchant(
                merchant, description, "wechat"
            )

            items.append({
                "order_no": row.get("交易单号", "").strip(),
                "transaction_time": row.get("交易时间", "").strip(),
                "merchant": detected_merchant,
                "description": description,
                "amount": int(amount * 100),
                "type": txn_type,
                "platform": detected_platform,
                "payment_method": payment_method,
            })
        except Exception:
            continue

    meta = {
        "platform": "微信",
        "has_payment_method": True,
        "detected_methods": list(methods),
    }
    return items, meta


def parse_excel(content: bytes) -> tuple[list[dict], dict]:
    """解析 Excel 格式账单（xlsx/xls），主要针对微信 Excel 账单

    Returns:
        (items, meta)
    """
    import openpyxl

    wb = openpyxl.load_workbook(content, read_only=True, data_only=True)
    ws = wb.active

    # 检测是否为微信账单（前10行包含"微信"）
    is_wechat = False
    for row in ws.iter_rows(max_row=10, values_only=True):
        if any(cell and "微信" in str(cell) for cell in row):
            is_wechat = True
            break

    # 查找表头行
    header_idx = -1
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_vals = [str(cell) if cell else "" for cell in row]
        row_str = " ".join(row_vals)
        if "交易时间" in row_str and ("交易对方" in row_str or "金额" in row_str):
            header_idx = i
            headers = row_vals
            break

    if header_idx == -1:
        raise ValueError("无法识别 Excel 账单格式")

    items = []
    methods = set()

    for i, row in enumerate(ws.iter_rows(min_row=header_idx + 2, values_only=True)):
        try:
            row_dict = {}
            for j, cell in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = cell

            # 金额处理
            amount_val = row_dict.get("金额(元)") or row_dict.get("金额")
            if amount_val is None:
                continue
            if isinstance(amount_val, str):
                amount_val = amount_val.replace(",", "").replace("¥", "").strip()
            amount = float(amount_val)

            # 收/支方向
            direction = str(row_dict.get("收/支", "")).strip()
            if "不计收支" in direction:
                continue
            txn_type = "expense" if "支出" in direction else "income"

            # 状态过滤
            status = str(row_dict.get("当前状态", "")).strip()
            if "退款" in status:
                continue

            # 支付方式
            payment_method = str(row_dict.get("支付方式", "")).strip()
            if payment_method:
                methods.add(payment_method)

            # 时间处理（datetime 对象转字符串）
            txn_time = row_dict.get("交易时间", "")
            if isinstance(txn_time, datetime):
                txn_time = txn_time.strftime("%Y-%m-%d %H:%M:%S")
            else:
                txn_time = str(txn_time).strip()

            # 交易号
            order_no = row_dict.get("交易单号") or row_dict.get("交易号") or ""

            # 智能识别平台和商户
            merchant = str(row_dict.get("交易对方", "")).strip()
            description = str(row_dict.get("商品", "")).strip()
            source = "wechat" if is_wechat else "unknown"
            detected_platform, detected_merchant = identify_platform_and_merchant(
                merchant, description, source
            )

            items.append({
                "order_no": str(order_no).strip(),
                "transaction_time": txn_time,
                "merchant": detected_merchant,
                "description": description,
                "amount": int(amount * 100),
                "type": txn_type,
                "platform": detected_platform,
                "payment_method": payment_method,
            })
        except Exception:
            continue

    meta = {
        "platform": "微信" if is_wechat else "未知",
        "has_payment_method": True,
        "detected_methods": list(methods),
    }
    return items, meta
