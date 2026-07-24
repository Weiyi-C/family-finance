import csv
import io
import json
from datetime import datetime

import structlog
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user
from app.database import get_db
from app.models.bill_import import BillImport, BillImportItem
from app.models.category import Category
from app.models.channel import PaymentChannel
from app.models.transaction import Transaction
from app.models.user import User

logger = structlog.get_logger()
router = APIRouter(tags=["账单导入"])


def detect_and_decode(content_bytes: bytes) -> str:
    """自动检测编码并解码"""
    for encoding in ["utf-8-sig", "utf-8", "gbk", "gb18030", "gb2312"]:
        try:
            return content_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return content_bytes.decode("utf-8", errors="ignore")


# 已知平台列表（用于识别"交易对方"中的平台）
KNOWN_PLATFORMS = {
    "淘宝", "淘宝闪购", "天猫", "京东", "拼多多", "美团", "饿了么",
    "抖音", "小红书", "得物", "唯品会", "闲鱼", "亚马逊",
    "苏宁易购", "当当", "网易严选", "小米商城", "华为商城",
}


def identify_platform_and_merchant(counterparty: str, description: str, source: str) -> tuple[str, str]:
    """识别平台和商户

    返回: (platform, merchant)
    """
    # 检查"交易对方"是否是已知平台
    for platform in KNOWN_PLATFORMS:
        if platform in counterparty:
            # 如果是平台，尝试从"商品名称"中提取真实商户
            # 格式通常是: "商户名-商品描述" 或 "商户名(分店)"
            merchant = counterparty  # 默认用平台名

            # 尝试从描述中提取商户
            if description:
                # 模式1: "商户名(分店)外卖订单"
                if "外卖订单" in description:
                    merchant = description.replace("外卖订单", "").strip()
                # 模式2: "商户名-商品描述"
                elif "-" in description:
                    parts = description.split("-", 1)
                    if len(parts[0]) > 2 and len(parts[0]) < 20:  # 合理的商户名长度
                        merchant = parts[0].strip()
                # 模式3: "商户名(分店)"
                elif "(" in description and ")" in description:
                    match = description[:description.index(")") + 1]
                    if len(match) > 2:
                        merchant = match.strip()

            return platform, merchant

    # 不是已知平台
    if source == "alipay":
        return "支付宝", counterparty
    elif source == "wechat":
        return "微信", counterparty
    else:
        return "线下", counterparty


def parse_alipay_csv(content: str) -> tuple[list[dict], dict]:
    """解析支付宝 CSV 账单

    返回: (交易列表, 元数据)
    元数据包含: platform(平台), has_payment_method(是否包含具体支付方式)
    """
    lines = content.strip().split("\n")
    header_idx = -1
    for i, line in enumerate(lines):
        if "交易" in line and ("对方" in line or "金额" in line):
            header_idx = i
            break
    if header_idx == -1:
        raise ValueError("无法识别支付宝账单格式")

    headers = [h.strip() for h in lines[header_idx].split(",") if h.strip()]
    reader = csv.DictReader(lines[header_idx:])
    items = []

    # 检查是否有"收/付款方式"列（新版支付宝可能有）
    has_payment_method = "收/付款方式" in headers or "付款方式" in headers

    for row in reader:
        try:
            # 金额
            amount_key = next((k for k in row if k and "金额" in k), None)
            if not amount_key or not row.get(amount_key):
                continue
            amount = float(row[amount_key].replace(",", "").replace("¥", "").strip())

            # 收/支
            dir_key = next((k for k in row if k and "收" in k and "支" in k), None)
            direction = row.get(dir_key, "").strip() if dir_key else ""
            if "不计收支" in direction:
                continue
            txn_type = "expense" if "支出" in direction else "income"

            # 状态 - 跳过退款
            status_key = next((k for k in row if k and "状态" in k), None)
            status_val = row.get(status_key, "").strip() if status_key else ""
            if "退款" in status_val:
                continue

            # 交易号
            order_key = next((k for k in row if k and "交易号" in k and "商家" not in k), None)
            order_no = row.get(order_key, "").strip() if order_key else ""

            # 交易时间
            time_key = next((k for k in row if k and ("创建时间" in k or "付款时间" in k or "交易时间" in k)), None)
            txn_time = row.get(time_key, "").strip() if time_key else ""

            # 交易对方(商户)
            merchant_key = next((k for k in row if k and "对方" in k), None)
            merchant = row.get(merchant_key, "").strip() if merchant_key else ""

            # 商品名称
            desc_key = next((k for k in row if k and ("商品" in k or "名称" in k)), None)
            description = row.get(desc_key, "").strip() if desc_key else ""

            # 交易来源地(平台) - 支付宝网站/APP等
            platform_key = next((k for k in row if k and "来源" in k), None)
            platform = row.get(platform_key, "").strip() if platform_key else ""

            # 收/付款方式(具体资金来源) - 如果有的话
            payment_key = next((k for k in row if k and ("付款方式" in k or "收/付款方式" in k)), None)
            payment_method = row.get(payment_key, "").strip() if payment_key else ""

            # 资金状态
            fund_key = next((k for k in row if k and "资金" in k), None)
            fund_status = row.get(fund_key, "").strip() if fund_key else ""

            # 类型(交易方式)
            type_key = next((k for k in row if k and k == "类型"), None)
            txn_method = row.get(type_key, "").strip() if type_key else ""

            # 尝试从描述推断资金来源
            inferred_account = ""
            if not payment_method:
                # 从商品名称或备注推断
                combined = f"{description} {merchant}"
                if "花呗" in combined:
                    inferred_account = "花呗"
                elif "余额宝" in combined:
                    inferred_account = "余额宝"
                elif "借呗" in combined:
                    inferred_account = "借呗"
                elif "信用卡" in combined:
                    inferred_account = "信用卡"

            # 智能识别平台和商户
            detected_platform, detected_merchant = identify_platform_and_merchant(
                merchant, description, "alipay"
            )

            items.append({
                "order_no": order_no,
                "transaction_time": txn_time,
                "merchant": detected_merchant,  # 使用识别后的商户
                "description": description,
                "amount": int(amount * 100),  # 转为分
                "type": txn_type,
                "platform": detected_platform,  # 使用识别后的平台
                "platform_raw": platform,  # 保留原始来源地
                "payment_method": payment_method or inferred_account,
                "fund_status": fund_status,
                "txn_method": txn_method,
            })
        except Exception as e:
            logger.warning("parse_alipay_row_error", error=str(e))

    meta = {
        "platform": "支付宝",
        "has_payment_method": has_payment_method,
        "detected_methods": list(set(i["payment_method"] for i in items if i.get("payment_method"))),
    }
    return items, meta


def parse_wechat_csv(content: str) -> tuple[list[dict], dict]:
    """解析微信 CSV 账单"""
    lines = content.strip().split("\n")
    header_idx = -1
    for i, line in enumerate(lines):
        if "交易时间" in line and "交易对方" in line:
            header_idx = i
            break
    if header_idx == -1:
        raise ValueError("无法识别微信账单格式")

    reader = csv.DictReader(lines[header_idx:])
    items = []
    methods = set()

    for row in reader:
        try:
            amount_str = row.get("金额(元)", "0").replace(",", "").replace("¥", "").strip()
            amount = float(amount_str)
            direction = row.get("收/支", "").strip()
            if "不计收支" in direction:
                continue
            txn_type = "expense" if "支出" in direction else "income"
            status = row.get("当前状态", "").strip()
            if "已退款" in status or "退款" in status:
                continue

            # 微信账单有明确的"支付方式"列
            payment_method = row.get("支付方式", "").strip()
            if payment_method:
                methods.add(payment_method)

            # 智能识别平台和商户
            merchant = row.get("交易对方", "").strip()
            description = row.get("商品", "").strip()
            detected_platform, detected_merchant = identify_platform_and_merchant(
                merchant, description, "wechat"
            )

            items.append({
                "order_no": row.get("交易单号", "").strip(),
                "transaction_time": row.get("交易时间", "").strip(),
                "merchant": detected_merchant,
                "description": description,
                "amount": int(amount * 100),
                "type": txn_type,
                "platform": detected_platform,
                "payment_method": payment_method,
            })
        except Exception as e:
            logger.warning("parse_wechat_row_error", error=str(e))

    meta = {
        "platform": "微信",
        "has_payment_method": True,  # 微信账单总是有支付方式
        "detected_methods": list(methods),
    }
    return items, meta


def parse_excel(content: bytes) -> tuple[list[dict], dict]:
    """解析 Excel 账单（支持微信等）"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器未安装 openpyxl")

    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {"platform": "unknown", "has_payment_method": False, "detected_methods": []}

    # 检测是否是微信账单
    first_few_rows = " ".join(str(c) for row in rows[:10] for c in row if c)
    is_wechat = "微信" in first_few_rows

    # 查找表头行
    header_idx = -1
    for i, row in enumerate(rows):
        row_str = " ".join(str(c) for c in row if c)
        if "交易时间" in row_str and ("对方" in row_str or "商品" in row_str):
            header_idx = i
            break
        # 微信Excel表头在"交易时间"行
        if row and row[0] and str(row[0]).strip() == "交易时间":
            header_idx = i
            break

    if header_idx == -1:
        wb.close()
        raise ValueError("无法识别 Excel 账单格式")

    headers = [str(c).strip() if c else "" for c in rows[header_idx]]
    items = []
    payment_methods = set()

    for row in rows[header_idx + 1:]:
        data = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                data[headers[j]] = val

        # 跳过空行
        if not any(data.values()):
            continue

        # 金额
        amount_key = next((k for k in data if "金额" in k), None)
        if not amount_key or not data.get(amount_key):
            continue

        try:
            amount_val = data.get(amount_key, 0)
            if isinstance(amount_val, str):
                amount_val = float(amount_val.replace(",", "").replace("¥", "").replace("元", ""))
            amount = float(amount_val)
            if amount <= 0:
                continue

            # 收/支
            dir_key = next((k for k in data if "收" in k and "支" in k), None)
            direction = str(data.get(dir_key, "")) if dir_key else ""
            if "不计收支" in direction:
                continue
            txn_type = "expense" if "支出" in direction else "income"

            # 状态 - 跳过退款
            status_key = next((k for k in data if "状态" in k), None)
            status_val = str(data.get(status_key, "")) if status_key else ""
            if "退款" in status_val:
                continue

            # 交易时间
            time_key = next((k for k in data if "时间" in k), None)
            txn_time = data.get(time_key, "")
            if hasattr(txn_time, 'strftime'):
                txn_time = txn_time.strftime("%Y-%m-%d %H:%M:%S")
            else:
                txn_time = str(txn_time) if txn_time else ""

            # 交易对方
            merchant_key = next((k for k in data if "对方" in k), None)
            merchant = str(data.get(merchant_key, "")).strip() if merchant_key else ""

            # 商品描述
            desc_key = next((k for k in data if "商品" in k or "名称" in k), None)
            description = str(data.get(desc_key, "")).strip() if desc_key else ""

            # 支付方式（微信账单特有）
            payment_key = next((k for k in data if "支付方式" in k), None)
            payment_method = str(data.get(payment_key, "")).strip() if payment_key else ""
            if payment_method:
                payment_methods.add(payment_method)

            # 交易号
            order_key = next((k for k in data if "交易单号" in k or "交易号" in k), None)
            order_no = str(data.get(order_key, "")).strip() if order_key else ""

            # 智能识别平台和商户
            platform = "微信" if is_wechat else ""
            detected_platform, detected_merchant = identify_platform_and_merchant(
                merchant, description, "wechat" if is_wechat else ""
            )

            items.append({
                "order_no": order_no,
                "transaction_time": txn_time,
                "merchant": detected_merchant,
                "description": description,
                "amount": int(amount * 100),
                "type": txn_type,
                "platform": detected_platform,
                "payment_method": payment_method,
            })
        except Exception as e:
            logger.warning("parse_excel_row_error", error=str(e))

    wb.close()

    return items, {
        "platform": "微信" if is_wechat else "unknown",
        "has_payment_method": len(payment_methods) > 0,
        "detected_methods": list(payment_methods),
    }


# 支付方式到账户类型的映射
PAYMENT_METHOD_TYPE_MAP = {
    "零钱": "wechat_balance",
    "零钱通": "wechat_lingqian",
    "工商银行": "bank_savings",
    "建设银行": "bank_savings",
    "农业银行": "bank_savings",
    "中国银行": "bank_savings",
    "交通银行": "bank_savings",
    "招商银行": "bank_savings",
    "储蓄卡": "bank_savings",
    "信用卡": "bank_credit",
}


def match_payment_method_to_account(method: str, accounts: list) -> int | None:
    """尝试将支付方式匹配到已有账户

    匹配规则：
    1. 精确匹配账户名称
    2. 匹配银行卡尾号（如"工商银行储蓄卡(0726)" → 匹配尾号0726的账户）
    3. 匹配关键词（如"零钱" → 微信零钱）
    """
    if not method or not accounts:
        return None

    method_clean = method.strip()

    # 1. 精确匹配
    for acc in accounts:
        if acc.name == method_clean:
            return acc.id

    # 2. 提取尾号匹配（格式：银行xxx(尾号)）
    import re
    tail_match = re.search(r'\((\d{4})\)', method_clean)
    if tail_match:
        tail = tail_match.group(1)
        for acc in accounts:
            if acc.card_tail == tail:
                return acc.id

    # 3. 关键词匹配
    for acc in accounts:
        # 零钱 → 微信零钱
        if method_clean == "零钱" and "零钱" in acc.name and "零钱通" not in acc.name:
            return acc.id
        # 零钱通
        if method_clean == "零钱通" and "零钱通" in acc.name:
            return acc.id
        # 银行卡匹配
        for bank_keyword in ["工商银行", "建设银行", "农业银行", "中国银行", "交通银行", "招商银行"]:
            if bank_keyword in method_clean and bank_keyword in acc.name:
                return acc.id

    return None


def suggest_account_type(method: str) -> dict:
    """根据支付方式建议账户类型"""
    method_clean = method.strip()

    if "零钱通" in method_clean:
        return {"type_code": "wechat_lingqian", "name": "零钱通", "group": "微信"}
    elif "零钱" in method_clean:
        return {"type_code": "wechat_balance", "name": "微信零钱", "group": "微信"}
    elif "储蓄卡" in method_clean or "借记卡" in method_clean:
        # 提取银行名
        import re
        bank_match = re.search(r'([\u4e00-\u9fa5]+银行)', method_clean)
        bank_name = bank_match.group(1) if bank_match else "银行"
        tail_match = re.search(r'\((\d{4})\)', method_clean)
        tail = tail_match.group(1) if tail_match else ""
        return {"type_code": "bank_savings", "name": f"{bank_name}储蓄卡", "bank_name": bank_name, "card_tail": tail, "group": bank_name}
    elif "信用卡" in method_clean:
        import re
        bank_match = re.search(r'([\u4e00-\u9fa5]+银行)', method_clean)
        bank_name = bank_match.group(1) if bank_match else "银行"
        tail_match = re.search(r'\((\d{4})\)', method_clean)
        tail = tail_match.group(1) if tail_match else ""
        return {"type_code": "bank_credit", "name": f"{bank_name}信用卡", "bank_name": bank_name, "card_tail": tail, "group": bank_name}
    else:
        return {"type_code": "e_wallet", "name": method_clean, "group": "其他"}


# ===================== 自动分类规则 =====================
# 按优先级排序，先匹配先生效
# category_expense: 支出分类名 | category_income: 收入分类名
AUTO_CATEGORY_RULES = [
    # -- 转账/红包（最高优先级，先排除） --
    {"keywords": ["红包"], "category_expense": "红包", "category_income": "红包收入", "tags": ["红包"]},
    {"keywords": ["转账", "还款", "代付", "亲属卡"], "category_expense": "转账", "category_income": "转账收入", "tags": ["转账"]},

    # -- 餐饮 --
    {"keywords": ["外卖", "美团外卖", "饿了么"], "category_expense": "餐饮", "tags": ["外卖"]},
    {"keywords": ["麦当劳", "肯德基", "KFC", "星巴克", "瑞幸", "喜茶", "奈雪", "蜜雪冰城", "必胜客", "汉堡王", "海底捞", "西贝"], "category_expense": "餐饮", "tags": ["餐饮"]},
    {"keywords": ["餐厅", "饭店", "食堂", "小吃", "烧烤", "火锅", "面馆", "早餐", "午餐", "晚餐", "下午茶", "咖啡", "奶茶", "蛋糕", "面包"], "category_expense": "餐饮", "tags": ["餐饮"]},
    {"keywords": ["美团", "饿了么"], "category_expense": "餐饮", "tags": ["外卖"]},

    # -- 食材采购 --
    {"keywords": ["超市", "便利店", "菜市场", "水果", "盒马", "永辉", "大润发", "沃尔玛", "家乐福", "山姆", "Costco", "开市客"], "category_expense": "食材采购", "tags": ["超市"]},

    # -- 交通出行 --
    {"keywords": ["滴滴", "打车", "出租车", "曹操出行", "T3出行", "高德打车"], "category_expense": "交通出行", "tags": ["打车"]},
    {"keywords": ["停车", "加油", "充电桩", "中石油", "中石化", "ETC"], "category_expense": "交通出行", "tags": ["交通"]},
    {"keywords": ["地铁", "公交", "一卡通", "交通卡"], "category_expense": "交通出行", "tags": ["交通"]},

    # -- 差旅 --
    {"keywords": ["火车", "高铁", "12306", "机票", "飞机", "携程", "去哪儿", "飞猪", "同程", "酒店", "民宿", "Airbnb"], "category_expense": "差旅", "tags": ["出差"]},

    # -- 购物 --
    {"keywords": ["淘宝", "天猫", "京东", "拼多多", "抖音商城", "抖音小店", "唯品会", "得物", "闲鱼"], "category_expense": "购物", "tags": ["网购"]},
    {"keywords": ["Apple", "苹果", "华为", "小米", "OPPO", "VIVO", "三星"], "category_expense": "购物", "tags": ["购物"]},

    # -- 休闲娱乐 --
    {"keywords": ["电影", "影院", "万达影城", "猫眼", "淘票票"], "category_expense": "休闲娱乐", "tags": ["娱乐"]},
    {"keywords": ["游戏", "Steam", "腾讯游戏", "网易游戏", "App Store", "Apple Music", "爱奇艺", "优酷", "腾讯视频", "B站", "bilibili", "Netflix", "Spotify"], "category_expense": "休闲娱乐", "tags": ["订阅"]},
    {"keywords": ["KTV", "酒吧", "网吧", "剧本杀", "密室"], "category_expense": "休闲娱乐", "tags": ["娱乐"]},

    # -- 居住 --
    {"keywords": ["电费", "水费", "燃气", "暖气", "物业", "房租", "房贷", "按揭"], "category_expense": "住房", "tags": ["居住"]},
    {"keywords": ["装修", "家具", "家电", "宜家"], "category_expense": "住房", "tags": ["购物"]},

    # -- 通讯 --
    {"keywords": ["话费", "流量", "宽带", "中国移动", "中国联通", "中国电信"], "category_expense": "通讯", "tags": ["通讯"]},

    # -- 医疗健康 --
    {"keywords": ["医院", "药店", "诊所", "口腔", "牙科", "体检", "药房", "大参林", "老百姓大药房"], "category_expense": "医疗健康", "tags": ["医疗"]},

    # -- 教育 --
    {"keywords": ["学费", "培训", "课程", "书店", "图书", "当当", "得到", "知乎", "网课"], "category_expense": "教育学习", "tags": ["教育"]},

    # -- 服饰美容 --
    {"keywords": ["理发", "美甲", "化妆品", "护肤", "ZARA", "H&M", "优衣库", "耐克", "阿迪"], "category_expense": "服饰美容", "tags": ["美容"]},

    # -- 孝敬家长 --
    {"keywords": ["孝敬", "爸妈", "父母", "长辈"], "category_expense": "孝敬家长", "tags": ["孝敬"]},

    # -- 保险 --
    {"keywords": ["保险", "社保", "公积金", "平安保险", "中国人寿"], "category_expense": "保险", "tags": ["保险"]},

    # -- 投资 --
    {"keywords": ["基金", "股票", "理财", "定期", "利息"], "category_expense": "投资亏损", "category_income": "投资收益", "tags": ["投资"]},
]

# 平台 → 默认分类映射（当关键词匹配失败时使用）
PLATFORM_CATEGORY_MAP = {
    "美团": "餐饮",
    "饿了么": "餐饮",
    "大众点评": "餐饮",
    "淘宝": "购物",
    "天猫": "购物",
    "京东": "购物",
    "拼多多": "购物",
    "抖音": "购物",
    "闲鱼": "购物",
    "携程": "差旅",
    "去哪儿": "差旅",
    "飞猪": "差旅",
    "滴滴": "交通出行",
}

# 支付方式 → 默认分类映射
PAYMENT_METHOD_CATEGORY_MAP = {
    "花呗": "购物",  # 花呗多数用于网购
    "余额宝": None,  # 余额宝是资金来源，不能推断分类
    "零钱通": None,
}

# 平台 → 默认标签
PLATFORM_TAG_MAP = {
    "淘宝": "网购",
    "天猫": "网购",
    "京东": "网购",
    "拼多多": "网购",
    "抖音": "网购",
    "美团": "外卖",
    "饿了么": "外卖",
    "携程": "出差",
    "去哪儿": "出差",
}


def auto_categorize(merchant: str, description: str, txn_type: str = "expense",
                    platform: str = "", payment_method: str = "") -> tuple[str | None, list[str]]:
    """自动分类 + 自动标签，返回 (分类名, 标签列表)"""
    combined = f"{merchant} {description}".lower()
    tags: list[str] = []

    # 1. 关键词规则匹配
    for rule in AUTO_CATEGORY_RULES:
        for keyword in rule["keywords"]:
            if keyword.lower() in combined:
                cat = rule.get("category_income") if txn_type == "income" and "category_income" in rule else rule["category_expense"]
                tags.extend(rule.get("tags", []))
                return cat, tags

    # 2. 平台分类回退
    if platform:
        cat = PLATFORM_CATEGORY_MAP.get(platform)
        if cat:
            tag = PLATFORM_TAG_MAP.get(platform)
            if tag:
                tags.append(tag)
            return cat, tags

    # 3. 支付方式回退
    if payment_method:
        cat = PAYMENT_METHOD_CATEGORY_MAP.get(payment_method)
        if cat:
            return cat, tags

    return None, tags


async def _ai_suggest_category(db: AsyncSession, user: User, merchant: str, description: str) -> str | None:
    """调用 AI 规则引擎推荐分类（与 ai.py 的 suggest_category 逻辑一致）"""
    from app.api.ai import CATEGORY_KEYWORDS
    combined = f"{merchant} {description}".lower()
    for cat_name, keywords in CATEGORY_KEYWORDS.items():
        for keyword in keywords:
            if keyword in combined:
                return cat_name
    return None


async def _auto_assign_tags(db: AsyncSession, family_id: int, txn_id: int, tag_names: list[str]):
    """自动创建并分配标签到交易"""
    from app.models.tag import Tag
    from app.models.transaction_tag import TransactionTag
    for name in tag_names:
        if not name:
            continue
        # 查找或创建标签
        result = await db.execute(
            select(Tag).where(Tag.family_id == family_id, Tag.name == name)
        )
        tag = result.scalar_one_or_none()
        if not tag:
            tag = Tag(family_id=family_id, name=name)
            db.add(tag)
            await db.flush()
        # 创建关联（忽略重复）
        existing = await db.execute(
            select(TransactionTag).where(
                TransactionTag.transaction_id == txn_id,
                TransactionTag.tag_id == tag.id,
            )
        )
        if not existing.scalar_one_or_none():
            db.add(TransactionTag(transaction_id=txn_id, tag_id=tag.id))


@router.post("/api/imports/upload")
async def upload_import(
    file: UploadFile = File(...),
    book_id: int = Form(...),
    source: str = Form("auto"),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """上传账单文件并解析"""
    content_bytes = await file.read()
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    try:
        if ext == "csv":
            content_str = detect_and_decode(content_bytes)
            if source == "auto":
                if "支付宝" in content_str[:500] or "交易号" in content_str[:1000]:
                    source = "alipay"
                elif "微信" in content_str[:500] or "交易单号" in content_str[:1000]:
                    source = "wechat"
                else:
                    source = "alipay"

            if source == "alipay":
                items, meta = parse_alipay_csv(content_str)
            elif source == "wechat":
                items, meta = parse_wechat_csv(content_str)
            else:
                items, meta = parse_alipay_csv(content_str)
            file_format = "csv"

        elif ext in ("xlsx", "xls"):
            items, meta = parse_excel(content_bytes)
            file_format = "xlsx"
        else:
            raise HTTPException(status_code=400, detail=f"不支持的文件格式: {ext}")

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("import_parse_error", error=str(e))
        raise HTTPException(status_code=500, detail=f"解析失败: {str(e)}")

    # 去重
    order_nos = [i["order_no"] for i in items if i.get("order_no")]
    existing_orders = set()
    if order_nos:
        result = await db.execute(
            select(BillImportItem.raw_data).where(
                BillImportItem.import_id.in_(
                    select(BillImport.id).where(BillImport.family_id == current_user.family_id)
                )
            )
        )
        for row in result.scalars():
            if row and row.get("order_no"):
                existing_orders.add(row["order_no"])

    new_items = [i for i in items if not i.get("order_no") or i["order_no"] not in existing_orders]
    skipped_dup = len(items) - len(new_items)

    # === 自动匹配账户 ===
    from app.models.account import PaymentAccount
    accounts_result = await db.execute(
        select(PaymentAccount).where(
            PaymentAccount.family_id == current_user.family_id,
            PaymentAccount.is_active == True,
        )
    )
    user_accounts = list(accounts_result.scalars())

    detected_methods = meta.get("detected_methods", [])
    method_matches = {}
    unmatched_methods = []
    for method in detected_methods:
        matched_id = match_payment_method_to_account(method, user_accounts)
        if matched_id:
            method_matches[method] = matched_id
        else:
            suggestion = suggest_account_type(method)
            unmatched_methods.append({"method": method, "suggestion": suggestion})

    # === 自动分类和标签（在存储之前完成） ===
    categories_result = await db.execute(
        select(Category).where(
            (Category.family_id == current_user.family_id) | (Category.family_id.is_(None)),
        )
    )
    category_names = {c.name: c.id for c in categories_result.scalars()}

    for item in new_items:
        txn_type = item.get("type", "expense")
        merchant = item.get("merchant", "")
        description = item.get("description", "")
        platform = item.get("platform", "")
        pm = item.get("payment_method", "")

        # 1. 规则分类 + 标签
        suggested_cat, auto_tags = auto_categorize(merchant, description, txn_type, platform, pm)

        # 2. AI 回退
        if not suggested_cat and (merchant or description):
            try:
                ai_result = await _ai_suggest_category(db, current_user, merchant, description)
                if ai_result:
                    suggested_cat = ai_result
            except Exception:
                pass

        # 3. 写入分类建议
        if suggested_cat:
            cat_id = category_names.get(suggested_cat)
            if cat_id:
                item["suggested_category_id"] = cat_id
                item["suggested_category_name"] = suggested_cat

        # 4. 写入标签建议
        if auto_tags:
            item["suggested_tags"] = auto_tags

        # 5. 自动匹配账户
        if pm and pm in method_matches:
            item["suggested_account_id"] = method_matches[pm]

    # === 存储到数据库（此时 raw_data 已包含分类和标签） ===
    imp = BillImport(
        family_id=current_user.family_id,
        book_id=book_id,
        source=source,
        file_format=file_format,
        total_rows=len(items),
        parsed_count=len(new_items),
        imported_by=current_user.id,
        status="parsed",
    )
    db.add(imp)
    await db.flush()

    for item_data in new_items:
        item = BillImportItem(
            import_id=imp.id,
            raw_data=item_data,
            parsed_amount=item_data.get("amount"),
            parsed_merchant=item_data.get("merchant"),
            action="pending",
        )
        db.add(item)

    await db.commit()
    await db.refresh(imp)

    return {
        "id": imp.id,
        "source": source,
        "file_format": file_format,
        "total_rows": len(items),
        "parsed_count": len(new_items),
        "skipped_duplicate": skipped_dup,
        "status": "parsed",
        "meta": meta,
        "method_matches": method_matches,
        "unmatched_methods": unmatched_methods,
        "preview": new_items[:30],
    }


@router.get("/api/imports")
async def list_imports(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BillImport)
        .where(BillImport.family_id == current_user.family_id)
        .order_by(BillImport.created_at.desc())
    )
    rows = result.scalars().all()
    return [
        {
            "id": i.id, "family_id": i.family_id, "book_id": i.book_id,
            "source": i.source, "file_format": i.file_format, "status": i.status,
            "total_rows": i.total_rows, "parsed_count": i.parsed_count,
            "matched_count": i.matched_count, "new_count": i.new_count,
            "created_at": i.created_at.isoformat() if i.created_at else None,
        }
        for i in rows
    ]


@router.get("/api/imports/{import_id}")
async def get_import(
    import_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BillImport).where(
            BillImport.id == import_id,
            BillImport.family_id == current_user.family_id,
        )
    )
    imp = result.scalar_one_or_none()
    if not imp:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入记录不存在")
    return imp


@router.get("/api/imports/{import_id}/items")
async def list_import_items(
    import_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BillImportItem).where(BillImportItem.import_id == import_id)
    )
    rows = result.scalars().all()
    return [
        {
            "id": i.id, "import_id": i.import_id, "raw_data": i.raw_data,
            "parsed_amount": i.parsed_amount, "parsed_merchant": i.parsed_merchant,
            "action": i.action, "matched_txn_id": i.matched_txn_id,
        }
        for i in rows
    ]


@router.post("/api/imports/{import_id}/confirm")
async def confirm_import(
    import_id: int,
    body: dict | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """确认导入

    body: {
        "default_account_id": 1,  // 默认账户ID（用于没有明确支付方式的交易）
        "method_account_map": {   // 支付方式→账户ID映射
            "花呗": 2,
            "工商银行信用卡": 3,
            "零钱": 4
        }
    }
    """
    from sqlalchemy import text

    if body is None:
        body = {}

    default_account_id = body.get("default_account_id")
    method_account_map = body.get("method_account_map", {})

    result = await db.execute(
        select(BillImport).where(
            BillImport.id == import_id,
            BillImport.family_id == current_user.family_id,
        )
    )
    imp = result.scalar_one_or_none()
    if not imp:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入记录不存在")
    if imp.status == "confirmed":
        raise HTTPException(status_code=400, detail="该导入已确认")

    items_result = await db.execute(
        select(BillImportItem).where(
            BillImportItem.import_id == import_id,
            BillImportItem.action == "pending",
        )
    )
    items = items_result.scalars().all()

    # 加载分类名称→ID映射（用于实时分类）
    from app.models.category import Category
    cats_result = await db.execute(
        select(Category).where(
            (Category.family_id == current_user.family_id) | (Category.family_id.is_(None)),
        )
    )
    _category_name_map = {c.name: c.id for c in cats_result.scalars()}

    seq_result = await db.execute(text("SELECT nextval('entry_id_seq')"))
    base_entry_id = seq_result.scalar()

    created = 0
    skipped = 0
    for item in items:
        raw = item.raw_data
        amount = item.parsed_amount or raw.get("amount", 0)

        if not amount or amount <= 0:
            skipped += 1
            item.action = "skipped"
            continue

        entry_id = base_entry_id + created

        # 解析交易时间
        txn_time_str = raw.get("transaction_time", "")
        txn_time = datetime.now()
        if isinstance(txn_time_str, str) and txn_time_str:
            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d"]:
                try:
                    txn_time = datetime.strptime(txn_time_str.strip(), fmt)
                    break
                except ValueError:
                    continue

        # 确定账户ID：优先用 method_account_map，否则用 default_account_id
        payment_method = raw.get("payment_method", "")
        account_id = None
        if payment_method and payment_method in method_account_map:
            account_id = method_account_map[payment_method]
        elif default_account_id:
            account_id = default_account_id

        # 查找平台ID
        platform_name = raw.get("platform", "")
        platform_id = None
        if platform_name:
            from app.models.platform import Platform
            # 先尝试精确匹配（系统预设或家庭自定义）
            platform_result = await db.execute(
                select(Platform.id).where(
                    Platform.name == platform_name,
                    (Platform.family_id.is_(None)) | (Platform.family_id == current_user.family_id),
                ).limit(1)
            )
            platform_id = platform_result.scalar()
            # 如果没找到，用已知映射
            if not platform_id:
                PLATFORM_ALIASES = {
                    "淘宝闪购": "淘宝", "天猫": "淘宝", "淘宝商城": "淘宝",
                    "京东到家": "京东", "京东商城": "京东",
                    "美团外卖": "美团", "大众点评": "美团",
                    "饿了么星选": "饿了么",
                    "拼多多": "拼多多", "抖音商城": "抖音",
                }
                mapped_name = PLATFORM_ALIASES.get(platform_name)
                if mapped_name:
                    platform_result = await db.execute(
                        select(Platform.id).where(Platform.name == mapped_name).limit(1)
                    )
                    platform_id = platform_result.scalar()

        # 查找支付渠道ID（根据平台或来源推断）
        channel_id = None
        if imp.source == "alipay":
            channel_result = await db.execute(
                select(PaymentChannel.id).where(PaymentChannel.name == "支付宝").limit(1)
            )
            channel_id = channel_result.scalar()
        elif imp.source == "wechat":
            channel_result = await db.execute(
                select(PaymentChannel.id).where(PaymentChannel.name == "微信支付").limit(1)
            )
            channel_id = channel_result.scalar()

        txn_type = raw.get("type", "expense")

        # 使用自动分类建议（如果有的话），否则实时分类
        category_id = raw.get("suggested_category_id")
        suggested_tags = raw.get("suggested_tags", [])

        if not category_id:
            # raw_data 中没有分类（可能是旧代码上传的），实时分类
            merchant = raw.get("merchant", "")
            description = raw.get("description", "")
            platform = raw.get("platform", "")
            pm = raw.get("payment_method", "")
            cat_name, auto_tags = auto_categorize(merchant, description, txn_type, platform, pm)

            # AI 回退
            if not cat_name and (merchant or description):
                try:
                    cat_name = await _ai_suggest_category(db, current_user, merchant, description)
                except Exception:
                    pass

            if cat_name:
                cat_id = _category_name_map.get(cat_name)
                if cat_id:
                    category_id = cat_id
            if auto_tags:
                suggested_tags = auto_tags

        # 双式记账：创建 debit（业务信息）+ credit（资金来源）两条记录
        # 两边都设置 payment_account_id，与 create_transaction 保持一致
        debit = Transaction(
            family_id=current_user.family_id,
            book_id=imp.book_id,
            entry_id=entry_id,
            entry_side="debit",
            type=txn_type,
            amount=amount,
            currency="CNY",
            category_id=category_id,
            merchant_name=item.parsed_merchant or raw.get("merchant"),
            description=raw.get("description"),
            transaction_time=txn_time,
            payment_account_id=account_id,
            payment_channel_id=channel_id,
            platform_id=platform_id,
            recorded_by=current_user.id,
            paid_by=current_user.id,
            completion_status="complete",
            import_id=imp.id,
            raw_data=raw,
        )
        credit = Transaction(
            family_id=current_user.family_id,
            book_id=imp.book_id,
            entry_id=entry_id,
            entry_side="credit",
            type=txn_type,
            amount=amount,
            currency="CNY",
            payment_account_id=account_id,
            transaction_time=txn_time,
            recorded_by=current_user.id,
            import_id=imp.id,
        )
        db.add(debit)
        db.add(credit)
        await db.flush()

        # 自动打标签（使用 entry_id，与查询保持一致）
        if suggested_tags:
            await _auto_assign_tags(db, current_user.family_id, entry_id, suggested_tags)

        item.action = "imported"
        item.matched_txn_id = entry_id
        created += 1

    imp.status = "confirmed"
    imp.new_count = created
    await db.commit()

    msg = f"成功导入 {created} 条交易"
    if skipped:
        msg += f"，跳过 {skipped} 条(金额为0)"
    return {"message": msg, "imported": created, "skipped": skipped}


@router.delete("/api/imports/{import_id}", status_code=204)
async def delete_import(
    import_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(BillImport).where(
            BillImport.id == import_id,
            BillImport.family_id == current_user.family_id,
        )
    )
    imp = result.scalar_one_or_none()
    if not imp:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="导入记录不存在")

    # 先删除关联的交易记录（debit + credit 两行）
    from sqlalchemy import delete as sql_delete
    from app.models.transaction import Transaction
    await db.execute(
        sql_delete(Transaction).where(
            Transaction.import_id == import_id,
            Transaction.family_id == current_user.family_id,
        )
    )

    # 再删除关联的明细记录
    await db.execute(
        sql_delete(BillImportItem).where(BillImportItem.import_id == import_id)
    )

    await db.delete(imp)
    await db.commit()
